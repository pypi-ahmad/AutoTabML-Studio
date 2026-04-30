"""Excel workbook loader."""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any

import pandas as pd

from app.ingestion.base import BaseLoader
from app.ingestion.errors import ParseFailureError, UnsupportedSourceError
from app.ingestion.schemas import DatasetInputSpec
from app.ingestion.types import IngestionSourceType
from app.ingestion.url_loader import (
    async_fetch_url_to_temp_file,
    ensure_local_file_size_guard,
    fetch_url_to_temp_file,
)

_OPENPYXL_READ_ONLY_SUFFIXES = {".xlsx", ".xlsm"}
_OPENPYXL_READ_ONLY_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
}


class ExcelLoader(BaseLoader):
    """Load local or remote Excel workbooks."""

    supported_source_types = (IngestionSourceType.EXCEL,)

    def load_raw_dataframe(
        self,
        input_spec: DatasetInputSpec,
        *,
        row_limit: int | None = None,
    ) -> tuple[pd.DataFrame, dict[str, Any]]:
        try:
            if input_spec.path is not None:
                path = Path(input_spec.path)
                if not path.exists():
                    raise UnsupportedSourceError(f"Local Excel file not found: {path}")
                file_size = ensure_local_file_size_guard(
                    path,
                    max_bytes=input_spec.local_max_file_bytes,
                )
                content_type = None
                final_url = None
                source_kind = "path"
                dataframe, sheet_names, actual_sheet_name, used_read_only = self._read_excel_path(
                    path,
                    input_spec=input_spec,
                    row_limit=row_limit,
                    content_type=content_type,
                )
            elif input_spec.url is not None:
                with fetch_url_to_temp_file(
                    input_spec.url,
                    timeout=input_spec.url_timeout_seconds or 30.0,
                    max_download_bytes=input_spec.url_max_download_bytes,
                    max_redirects=input_spec.url_max_redirects,
                    max_retries=input_spec.url_max_retries,
                ) as downloaded:
                    file_size = downloaded.file_size_bytes
                    content_type = downloaded.content_type
                    final_url = downloaded.final_url
                    source_kind = "url"
                    dataframe, sheet_names, actual_sheet_name, used_read_only = self._read_excel_path(
                        downloaded.path,
                        input_spec=input_spec,
                        row_limit=row_limit,
                        content_type=content_type,
                    )
            else:
                raise UnsupportedSourceError("Excel ingestion requires either a local path or a URL.")

            source_details = {
                "source_kind": source_kind,
                "available_sheet_names": sheet_names,
                "sheet_name": actual_sheet_name,
                "content_type": content_type,
                "final_url": final_url,
                "file_size_bytes": file_size,
                "load_strategy": "openpyxl_read_only" if used_read_only else "pandas_read_excel",
            }
            return dataframe, source_details
        except UnsupportedSourceError:
            raise
        except IndexError as exc:
            raise ParseFailureError("The requested Excel sheet index is out of range.") from exc
        except KeyError as exc:
            raise ParseFailureError(f"The requested Excel sheet '{input_spec.excel_sheet}' was not found.") from exc
        except ValueError as exc:
            raise ParseFailureError(f"Failed to load the Excel workbook: {exc}") from exc
        except ImportError as exc:
            raise ParseFailureError(
                "Excel support requires the appropriate engine dependencies. Install 'openpyxl' and 'xlrd'."
            ) from exc

    async def load_raw_dataframe_async(
        self,
        input_spec: DatasetInputSpec,
        *,
        row_limit: int | None = None,
    ) -> tuple[pd.DataFrame, dict[str, Any]]:
        if input_spec.path is not None:
            return await asyncio.to_thread(
                self.load_raw_dataframe,
                input_spec,
                row_limit=row_limit,
            )

        try:
            if input_spec.url is not None:
                async with async_fetch_url_to_temp_file(
                    input_spec.url,
                    timeout=input_spec.url_timeout_seconds or 30.0,
                    max_download_bytes=input_spec.url_max_download_bytes,
                    max_redirects=input_spec.url_max_redirects,
                    max_retries=input_spec.url_max_retries,
                ) as downloaded:
                    dataframe, sheet_names, actual_sheet_name, used_read_only = await asyncio.to_thread(
                        self._read_excel_path,
                        downloaded.path,
                        input_spec=input_spec,
                        row_limit=row_limit,
                        content_type=downloaded.content_type,
                    )
                    source_details = {
                        "source_kind": "url",
                        "available_sheet_names": sheet_names,
                        "sheet_name": actual_sheet_name,
                        "content_type": downloaded.content_type,
                        "final_url": downloaded.final_url,
                        "file_size_bytes": downloaded.file_size_bytes,
                        "load_strategy": "openpyxl_read_only_async" if used_read_only else "pandas_read_excel_async",
                    }
                    return dataframe, source_details

            raise UnsupportedSourceError("Excel ingestion requires either a local path or a URL.")
        except UnsupportedSourceError:
            raise
        except IndexError as exc:
            raise ParseFailureError("The requested Excel sheet index is out of range.") from exc
        except KeyError as exc:
            raise ParseFailureError(f"The requested Excel sheet '{input_spec.excel_sheet}' was not found.") from exc
        except ValueError as exc:
            raise ParseFailureError(f"Failed to load the Excel workbook: {exc}") from exc
        except ImportError as exc:
            raise ParseFailureError(
                "Excel support requires the appropriate engine dependencies. Install 'openpyxl' and 'xlrd'."
            ) from exc

    def _resolve_sheet_name(self, sheet_name: str | int, sheet_names: list[str]) -> str | int:
        if isinstance(sheet_name, int):
            return sheet_names[sheet_name]
        return sheet_name

    def _read_excel_path(
        self,
        path: Path,
        *,
        input_spec: DatasetInputSpec,
        row_limit: int | None,
        content_type: str | None,
    ) -> tuple[pd.DataFrame, list[str], str | int, bool]:
        sheet_name = input_spec.excel_sheet if input_spec.excel_sheet is not None else 0
        if self._supports_read_only(path, content_type=content_type):
            dataframe, sheet_names, actual_sheet_name = self._read_excel_read_only(
                path,
                sheet_name=sheet_name,
                row_limit=row_limit,
            )
            return dataframe, sheet_names, actual_sheet_name, True

        with pd.ExcelFile(str(path)) as excel_file:
            sheet_names = list(excel_file.sheet_names)
            dataframe = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=row_limit)
            actual_sheet_name = self._resolve_sheet_name(sheet_name, sheet_names)
        return dataframe, sheet_names, actual_sheet_name, False

    def _supports_read_only(self, path: Path, *, content_type: str | None) -> bool:
        return (
            path.suffix.lower() in _OPENPYXL_READ_ONLY_SUFFIXES
            or content_type in _OPENPYXL_READ_ONLY_CONTENT_TYPES
        )

    def _read_excel_read_only(
        self,
        path: Path,
        *,
        sheet_name: str | int,
        row_limit: int | None,
    ) -> tuple[pd.DataFrame, list[str], str | int]:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
            actual_sheet_name = self._resolve_sheet_name(sheet_name, sheet_names)
            worksheet = workbook[actual_sheet_name]
            row_iter = worksheet.iter_rows(values_only=True)

            try:
                header_row = next(row_iter)
            except StopIteration:
                return pd.DataFrame(), sheet_names, actual_sheet_name

            headers = [
                value if value not in (None, "") else f"Unnamed: {index}"
                for index, value in enumerate(header_row)
            ]

            rows: list[tuple[Any, ...]] = []
            for values in row_iter:
                rows.append(tuple(values))
                if row_limit is not None and len(rows) >= row_limit:
                    break

            return pd.DataFrame(rows, columns=headers), sheet_names, actual_sheet_name
        finally:
            workbook.close()
