"""Unit tests for tabular ingestion loaders and routing."""

from __future__ import annotations

from pathlib import Path

import httpx
import pandas as pd
import pytest
import respx

from app.ingestion.errors import EmptyDatasetError, ParseFailureError, RemoteAccessError, UnsupportedSourceError
from app.ingestion.excel_loader import ExcelLoader
from app.ingestion.csv_loader import CSVLoader
from app.ingestion.factory import load_dataset, preview_dataset
from app.ingestion.schemas import DatasetInputSpec
from app.ingestion.types import IngestionSourceType


class TestLocalFileLoaders:
    def test_load_local_csv(self, tmp_path: Path):
        csv_path = tmp_path / "sample.csv"
        csv_path.write_text("feature,target\n1,0\n2,1\n", encoding="utf-8")

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.CSV, path=csv_path)
        )

        assert loaded.dataframe.shape == (2, 2)
        assert loaded.dataframe.columns.tolist() == ["feature", "target"]
        assert loaded.metadata.source_type == IngestionSourceType.CSV
        assert loaded.metadata.row_count == 2
        assert loaded.metadata.column_count == 2

    def test_load_local_delimited_text(self, tmp_path: Path):
        tsv_path = tmp_path / "sample.tsv"
        tsv_path.write_text("col_a\tcol_b\n10\t20\n30\t40\n", encoding="utf-8")

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.DELIMITED_TEXT, path=tsv_path)
        )

        assert loaded.dataframe.iloc[1].tolist() == [30, 40]
        assert loaded.metadata.source_details["delimiter"] == "\t"

    def test_load_excel_uses_first_sheet_by_default(self, tmp_path: Path):
        excel_path = tmp_path / "workbook.xlsx"
        with pd.ExcelWriter(excel_path) as writer:
            pd.DataFrame({"a": [1, 2]}).to_excel(writer, sheet_name="first", index=False)
            pd.DataFrame({"b": [3, 4]}).to_excel(writer, sheet_name="second", index=False)

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.EXCEL, path=excel_path)
        )

        assert loaded.dataframe.columns.tolist() == ["a"]
        assert loaded.metadata.sheet_name == "first"
        assert loaded.metadata.source_details["available_sheet_names"] == ["first", "second"]

    def test_local_csv_respects_file_size_guard(self, tmp_path: Path):
        csv_path = tmp_path / "sample.csv"
        csv_path.write_text("feature,target\n1,0\n2,1\n", encoding="utf-8")

        with pytest.raises(UnsupportedSourceError, match="safe read cap"):
            load_dataset(
                DatasetInputSpec(
                    source_type=IngestionSourceType.CSV,
                    path=csv_path,
                    local_max_file_bytes=8,
                )
            )

    def test_excel_loader_uses_openpyxl_read_only_mode(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
        excel_path = tmp_path / "workbook.xlsx"
        with pd.ExcelWriter(excel_path) as writer:
            pd.DataFrame({"a": [1, 2], "b": [3, 4]}).to_excel(writer, index=False)

        import openpyxl

        real_load_workbook = openpyxl.load_workbook
        captured: dict[str, Any] = {}

        def wrapped_load_workbook(*args, **kwargs):
            captured.update(kwargs)
            return real_load_workbook(*args, **kwargs)

        monkeypatch.setattr(openpyxl, "load_workbook", wrapped_load_workbook)

        loaded = ExcelLoader().load_raw_dataframe(
            DatasetInputSpec(source_type=IngestionSourceType.EXCEL, path=excel_path)
        )

        assert loaded[0].shape == (2, 2)
        assert captured["read_only"] is True
        assert captured["data_only"] is True


class TestRemoteLoaders:
    @respx.mock
    def test_load_html_table_from_url(self):
        url = "https://example.com/table"
        html = """
        <html><body>
        <table>
            <thead><tr><th>name</th><th>score</th></tr></thead>
            <tbody><tr><td>a</td><td>10</td></tr><tr><td>b</td><td>20</td></tr></tbody>
        </table>
        </body></html>
        """
        respx.get(url).mock(return_value=httpx.Response(200, text=html, headers={"content-type": "text/html"}))

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.HTML_TABLE, url=url)
        )

        assert loaded.dataframe.shape == (2, 2)
        assert loaded.metadata.source_details["detected_table_count"] == 1
        assert loaded.metadata.source_type == IngestionSourceType.HTML_TABLE
        assert loaded.metadata.source_details["load_strategy"] == "streamed_temp_file_bounded_html_parse"

    @respx.mock
    def test_load_html_table_from_url_respects_row_limit(self):
        url = "https://example.com/table-limited"
        html = """
        <html><body>
        <table>
            <thead><tr><th>name</th><th>score</th></tr></thead>
            <tbody>
                <tr><td>a</td><td>10</td></tr>
                <tr><td>b</td><td>20</td></tr>
                <tr><td>c</td><td>30</td></tr>
            </tbody>
        </table>
        </body></html>
        """
        respx.get(url).mock(return_value=httpx.Response(200, text=html, headers={"content-type": "text/html"}))

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.HTML_TABLE, url=url, row_limit=1)
        )

        assert loaded.dataframe.shape == (1, 2)
        assert loaded.dataframe.iloc[0].to_dict() == {"name": "a", "score": 10}
        assert loaded.metadata.applied_row_limit == 1

    @respx.mock
    def test_load_html_table_match_text_filters_tables(self):
        url = "https://example.com/table-match"
        html = """
        <html><body>
        <table>
            <tr><th>name</th></tr>
            <tr><td>alpha</td></tr>
        </table>
        <table>
            <caption>scores</caption>
            <tr><th>name</th><th>score</th></tr>
            <tr><td>beta</td><td>20</td></tr>
        </table>
        </body></html>
        """
        respx.get(url).mock(return_value=httpx.Response(200, text=html, headers={"content-type": "text/html"}))

        loaded = load_dataset(
            DatasetInputSpec(
                source_type=IngestionSourceType.HTML_TABLE,
                url=url,
                html_match_text="scores",
            )
        )

        assert loaded.dataframe.shape == (1, 2)
        assert loaded.metadata.source_details["detected_table_count"] == 1
        assert loaded.dataframe.iloc[0].to_dict() == {"name": "beta", "score": 20}

    @respx.mock
    def test_url_file_routes_to_csv_loader(self):
        url = "https://example.com/data"
        csv_body = b"x,y\n1,2\n3,4\n"
        respx.head(url).mock(return_value=httpx.Response(200, headers={"content-type": "text/csv"}))
        respx.get(url).mock(return_value=httpx.Response(200, content=csv_body, headers={"content-type": "text/csv"}))

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.URL_FILE, url=url)
        )

        assert loaded.dataframe.shape == (2, 2)
        assert loaded.metadata.source_type == IngestionSourceType.URL_FILE
        assert loaded.metadata.source_details["routed_source_type"] == "csv"

    @respx.mock
    def test_load_remote_excel_file(self):
        url = "https://example.com/dataset.xlsx"
        buffer = pd.io.common.BytesIO()
        with pd.ExcelWriter(buffer) as writer:
            pd.DataFrame({"age": [10, 20], "target": [0, 1]}).to_excel(writer, index=False)
        payload = buffer.getvalue()

        respx.get(url).mock(
            return_value=httpx.Response(
                200,
                content=payload,
                headers={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                },
            )
        )

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.EXCEL, url=url)
        )

        assert loaded.dataframe.shape == (2, 2)
        assert loaded.metadata.sheet_name == 0 or loaded.metadata.sheet_name == "Sheet1"

    @respx.mock
    def test_url_file_routes_to_html_table_via_sniff(self):
        url = "https://example.com/download?id=42"
        html = "<html><body><table><tr><th>a</th></tr><tr><td>1</td></tr></table></body></html>"
        respx.head(url).mock(return_value=httpx.Response(405))
        respx.get(url).mock(return_value=httpx.Response(200, text=html, headers={"content-type": "text/plain"}))

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.URL_FILE, url=url)
        )

        assert loaded.dataframe.iloc[0, 0] == 1
        assert loaded.metadata.source_details["routed_source_type"] == "html_table"
        assert loaded.metadata.source_details["probe_method"] == "get-sniff"

    @respx.mock
    def test_url_file_respects_configured_download_cap(self):
        url = "https://example.com/large.csv"
        csv_body = b"x,y\n1,2\n3,4\n"
        respx.head(url).mock(return_value=httpx.Response(200, headers={"content-type": "text/csv"}))
        respx.get(url).mock(return_value=httpx.Response(200, content=csv_body, headers={"content-type": "text/csv"}))

        with pytest.raises(RemoteAccessError, match="cap"):
            load_dataset(
                DatasetInputSpec(
                    source_type=IngestionSourceType.URL_FILE,
                    url=url,
                    url_max_download_bytes=8,
                )
            )


class TestDataFrameLoaderAndPreview:
    def test_dataframe_input_is_copied_and_normalized(self):
        original = pd.DataFrame(
            [[1, 2, None], [None, None, None]],
            columns=["feature", "feature", "empty"],
        )

        loaded = load_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.DATAFRAME, dataframe=original)
        )

        assert original.columns.tolist() == ["feature", "feature", "empty"]
        assert loaded.dataframe.columns.tolist() == ["feature", "feature__2"]
        assert loaded.metadata.normalization_actions
        assert loaded.dataframe.shape == (1, 2)

    def test_preview_dataset_limits_rows(self, tmp_path: Path):
        csv_path = tmp_path / "preview.csv"
        csv_path.write_text("a,b\n1,2\n3,4\n5,6\n", encoding="utf-8")

        loaded = preview_dataset(
            DatasetInputSpec(source_type=IngestionSourceType.CSV, path=csv_path),
            rows=2,
        )

        assert loaded.dataframe.shape == (2, 2)
        assert loaded.metadata.is_preview is True
        assert loaded.metadata.applied_row_limit == 2
        assert loaded.preview(1).shape == (1, 2)


class TestChunkedCSVReading:
    def test_csv_loader_reads_in_chunks_and_stops_after_row_limit(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ):
        csv_path = tmp_path / "chunked.csv"
        csv_path.write_text("feature,target\n1,0\n2,1\n3,0\n4,1\n5,0\n", encoding="utf-8")

        seen_kwargs: dict[str, Any] = {}
        consumed_chunks: list[int] = []

        class FakeReader:
            def __init__(self) -> None:
                self._chunks = iter(
                    [
                        pd.DataFrame({"feature": [1, 2], "target": [0, 1]}),
                        pd.DataFrame({"feature": [3, 4], "target": [0, 1]}),
                        pd.DataFrame({"feature": [5], "target": [0]}),
                    ]
                )

            def __iter__(self):
                return self

            def __next__(self):
                chunk = next(self._chunks)
                consumed_chunks.append(len(chunk))
                return chunk

        def fake_read_csv(*args, **kwargs):
            seen_kwargs.update(kwargs)
            return FakeReader()

        monkeypatch.setattr(pd, "read_csv", fake_read_csv)

        dataframe, source_details = CSVLoader().load_raw_dataframe(
            DatasetInputSpec(source_type=IngestionSourceType.CSV, path=csv_path),
            row_limit=3,
        )

        assert dataframe.shape == (3, 2)
        assert consumed_chunks == [2, 2]
        assert seen_kwargs["chunksize"] == 3
        assert source_details["load_strategy"] == "chunked_read_csv"


class TestFailures:
    @respx.mock
    def test_html_table_no_tables_raises(self):
        url = "https://example.com/no-table"
        respx.get(url).mock(
            return_value=httpx.Response(200, text="<html><body>No table here</body></html>", headers={"content-type": "text/html"})
        )

        with pytest.raises(ParseFailureError, match="No HTML tables were found"):
            load_dataset(DatasetInputSpec(source_type=IngestionSourceType.HTML_TABLE, url=url))

    @respx.mock
    def test_ambiguous_url_requires_override(self):
        url = "https://example.com/blob"
        respx.head(url).mock(return_value=httpx.Response(200, headers={"content-type": "application/octet-stream"}))
        respx.get(url).mock(return_value=httpx.Response(200, content=b"opaque payload without structured table data"))

        with pytest.raises(UnsupportedSourceError, match="Provide an explicit source_type override"):
            load_dataset(DatasetInputSpec(source_type=IngestionSourceType.URL_FILE, url=url))

    def test_empty_dataframe_raises(self):
        empty = pd.DataFrame({"a": [None], "b": [None]})
        with pytest.raises(EmptyDatasetError, match="empty after safe normalization"):
            load_dataset(DatasetInputSpec(source_type=IngestionSourceType.DATAFRAME, dataframe=empty))

    def test_missing_kaggle_package_is_actionable(self):
        with pytest.raises(
            RemoteAccessError,
            match="optional 'kaggle' package|Kaggle credentials are not configured",
        ):
            load_dataset(
                DatasetInputSpec(
                    source_type=IngestionSourceType.KAGGLE,
                    kaggle_dataset_ref="owner/dataset",
                )
            )
